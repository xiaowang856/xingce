import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE = ROOT.parent / "公务员行测申论复习计划-成语扩充版.xlsx"
OUTPUT = ROOT / "data.json"


def sheet_to_records(ws):
    headers = [cell.value for cell in ws[1]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        record = {}
        for header, value in zip(headers, row):
            if header is None:
                continue
            if hasattr(value, "isoformat"):
                value = value.isoformat()
            record[str(header)] = "" if value is None else value
        records.append(record)
    return records


def main():
    wb = load_workbook(SOURCE, data_only=False)
    data = {
        "weeklyPlan": sheet_to_records(wb["周计划"]),
        "dailyCheckin": sheet_to_records(wb["每日打卡"]),
        "examMistakes": sheet_to_records(wb["行测错题本"]),
        "shenlunPractice": sheet_to_records(wb["申论练习记录"]),
        "shenlunSites": sheet_to_records(wb["申论素材网站"]),
        "xingceSites": sheet_to_records(wb["行测练习网站"]),
        "focusList": sheet_to_records(wb["复习重点清单"]),
        "idioms": sheet_to_records(wb["成语积累"]),
    }
    OUTPUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(OUTPUT)


if __name__ == "__main__":
    main()
